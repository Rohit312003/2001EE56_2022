
# inmporting nessesry libreries
from collections import OrderedDict
from ast import Try
from operator import concat
import numpy as np
import pandas as pd
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook

#load excel file
try:
     book = load_workbook(filename="octant_input.xlsx")
except:
     print("unable to find the input file")
     exit()
 
#open workbook
sheet = book.active
 
#definging the property of border
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# ------------------------------------------------taking input file---------------------------------------------------------------
# appling try and except for  checking the file 
try:
                                                                      #reading the input file and storing its value in the df 
     df = pd.read_excel('octant_input.xlsx')
except:
     print("unable to find the input file")
     exit()

print("Column headings:")
print(df)
# ----------------------------------------------calculating the value of mean of column U,V and W-----------------------------------
try:
     meanU=df['U'].mean()
     meanV=df['V'].mean()
     meanW=df['W'].mean()
except:
     print("unable to find the mean")

sheet = book.active


meanU=df['U'].mean()
meanV=df['V'].mean()
meanW=df['W'].mean()
# adding the mean value to the new column
sheet.cell(row=1, column=5).value ="Uavg"
sheet.cell(row=2, column=5).value =meanU

sheet.cell(row=1, column=6).value ="Vavg"
sheet.cell(row=2, column=6).value =meanV

sheet.cell(row=1, column=7).value ="Wavg"
sheet.cell(row=2, column=7).value =meanW 


# saving the output to the output file
book.save("output_octant_longest_subsequence.xlsx")
# reading the same saved output value to new again and store it into df
df = pd.read_excel('output_octant_longest_subsequence.xlsx')
# ----------------------------------------------  calculating error in U,V and W----------------------------------
df['U-Uavg']=df['U']-meanU
df['V-Vavg']=df['V']-meanV
df['W-Wavg']=df['W']-meanW
# ----------------------------------------------below code for desiding the octant----------------------------------------------------------
                                                                  # creating a list for octant values
octant=[]                                                        
 #  calculating the Octant value .
try:
     for (l, m, q) in zip(df['U-Uavg'],df['V-Vavg'],df['W-Wavg']):     #   appanding the octannt values to the list
         if l>0 and m>0 and q<0:
             octant.append(-1)
         elif l>0 and m>0 and q>0:
              octant.append(1)
         elif l<0 and m>0 and q<0:
              octant.append(-2)
         elif l<0 and m>0 and q>0:
              octant.append(2)
         elif l<0 and m<0 and q<0:
              octant.append(-3)
         elif l<0 and m<0 and q>0:
           octant.append(3)
         elif l>=0 and m<=0 and q<0:
            octant.append(-4)
         elif l>0 and m<0 and q>0:
             octant.append(4) 
     df['octant'] = octant
except:
     print("somting wrong with the columns U-Uavg,V-Vavg,W-Wavg ")

# saving the file in out put file
with pd.ExcelWriter('output_octant_longest_subsequence.xlsx') as writer:
    df.to_excel(writer, sheet_name='Sheet_name_1',index=False)
print("sucssesfully done!")
# reading the same file again  in workbook

book = load_workbook(filename="output_octant_longest_subsequence.xlsx")
 
#open workbook
sheet = book.active
sheet.cell(row=4, column=12).value ="User input"
sheet.cell(row=4, column=12).border = thin_border

mod=5000
sheet = book.active
sheet.cell(row=4, column=13).value ="mod "+str(mod)
sheet.cell(row=4, column=13).border = thin_border

# making our desire matrix having desired heading



sheet.cell(row=2, column=13).value ="Octant ID"
sheet.cell(row=2, column=13).border = thin_border

sheet.cell(row=2, column=14).value =1
sheet.cell(row=2, column=14).border = thin_border

sheet.cell(row=2, column=15).value =-1
sheet.cell(row=2, column=15).border = thin_border

sheet.cell(row=2, column=16).value =2
sheet.cell(row=2, column=16).border = thin_border

sheet.cell(row=2, column=17).value =-2
sheet.cell(row=2, column=17).border = thin_border

sheet.cell(row=2, column=18).value =3
sheet.cell(row=2, column=18).border = thin_border

sheet.cell(row=2, column=19).value =-3
sheet.cell(row=2, column=19).border = thin_border

sheet.cell(row=2, column=20).value =4
sheet.cell(row=2, column=20).border = thin_border

sheet.cell(row=2, column=21).value =-4
sheet.cell(row=2, column=21).border = thin_border

sheet.cell(row=1, column=22).value =1
sheet.cell(row=1, column=22).border = thin_border
sheet.cell(row=2, column=22).value ="Rank 1"
sheet.cell(row=2, column=22).border = thin_border

sheet.cell(row=1, column=23).value =-1
sheet.cell(row=1, column=23).border = thin_border
sheet.cell(row=2, column=23).value ="Rank 2"
sheet.cell(row=2, column=23).border = thin_border

sheet.cell(row=1, column=24).value =2
sheet.cell(row=1, column=24).border = thin_border
sheet.cell(row=2, column=24).value ="Rank 3"
sheet.cell(row=2, column=24).border = thin_border

sheet.cell(row=1, column=25).value =-2
sheet.cell(row=1, column=25).border = thin_border
sheet.cell(row=2, column=25).value ="Rank 4"
sheet.cell(row=2, column=25).border = thin_border

sheet.cell(row=1, column=26).value =3
sheet.cell(row=1, column=26).border = thin_border
sheet.cell(row=2, column=26).value ="Rank 5"
sheet.cell(row=2, column=26).border = thin_border

sheet.cell(row=1, column=27).value =-3
sheet.cell(row=1, column=27).border = thin_border
sheet.cell(row=2, column=27).value ="Rank 6"
sheet.cell(row=2, column=27).border = thin_border

sheet.cell(row=1, column=28).value =4
sheet.cell(row=1, column=28).border = thin_border
sheet.cell(row=2, column=28).value ="Rank 7"
sheet.cell(row=2, column=28).border = thin_border

sheet.cell(row=1, column=29).value =-4
sheet.cell(row=1, column=29).border = thin_border
sheet.cell(row=2, column=29).value ="Rank 8"
sheet.cell(row=2, column=29).border = thin_border

sheet.cell(row=2, column=30).value ="Rank1 Octant ID"
sheet.cell(row=2, column=30).border = thin_border

sheet.cell(row=2, column=31).value ="Rank1 Octant Name"
sheet.cell(row=2, column=31).border = thin_border
# below code is for calculating the longest subsequence ans its count
sheet.cell(row=3, column=13).value ="Overall Count"
sheet.cell(row=3, column=13).border = thin_border


dict = {}
for k in range(1,5):
    maxo=0
    maxi=0
    for i in range(14,22,2):
    
        if(sheet.cell(row=2, column=i).value==k):
            sheet.cell(row=3, column=i).value =df['octant'].value_counts()[k]
            sheet.cell(row=3, column=i).border = thin_border
            dict[df['octant'].value_counts()[k]]=k
    for i in range(15,22,2):
    
        if(sheet.cell(row=2, column=i).value==-1*k):
            sheet.cell(row=3, column=i).value =df['octant'].value_counts()[-1*k]
            sheet.cell(row=3, column=i).border = thin_border
            dict[df['octant'].value_counts()[-1*k]]=-1*k


        



# Creates a sorted dictionary (sorted by key)

 

print(dict)
 
dict1 = OrderedDict(sorted(dict.items()))
k=8
q=0
dict2 ={}
for key in dict1:

  dict2[dict1[key]]=k
  k=k-1
for k in range(1,5):
    for i in range(23,30,2):
         if(sheet.cell(row=1, column=i).value==-1*k):
            sheet.cell(row=3, column=i).value =dict2[-1*k]
            sheet.cell(row=3, column=i).border = thin_border

    for i in range(22,30,2):
        if(sheet.cell(row=1, column=i).value==k):
            sheet.cell(row=3, column=i).value =dict2[k]
            sheet.cell(row=3, column=i).border = thin_border
    if(dict2[k]==1):
        q=k 
    if(dict2[-1*k]==1):
        q=k 

dict0={1:"Internal outward interaction",-1:"External outward interaction",2:"External Ejection",-2:"Internal Ejection",3:"External inward interaction",-3:"Internal inward interaction",4:"Internal sweep",-4:"External sweep"}
sheet.cell(row=3, column=30).value =q
sheet.cell(row=3, column=30).border = thin_border
sheet.cell(row=3, column=31).value =dict0[q]
sheet.cell(row=3, column=31).border = thin_border
  

leng=df['octant'].count()
print(leng) 
# Initialize matrix

 

k=mod
# printing the mod value






z=int(leng/k)+1

# no rows in output
R = int(z)

# no column is constant equal to 9
C = int(22)
  
# Initialize matrix

numb=[1,-1,2,-2,3,-3,4,-4]
  

zz=int(0)
dict5={"Internal outward interaction":0,"External outward interaction":0,"External Ejection":0,"Internal Ejection":0,"External inward interaction":0,"Internal inward interaction":0,"Internal sweep":0,"External sweep":0}
for i in range(R):
               # A for loop for row entries
    # creating a list
  
    dict = {}
    for j in range(13,C): 
     if j==13:
        # condition for last row of output to cheak the  length of input so it stop when input length reached
         if i==R-1:
            u=str(i*k)+"-"+str(leng) 
            sheet.cell(row=i+5, column=j).value =u
            sheet.cell(row=i+5, column=j).border = thin_border
         else:         
              u=str(i*k)+"-"+str(k*(i+1)-1) 
              sheet.cell(row=i+5, column=j).value =u
              sheet.cell(row=i+5, column=j).border = thin_border
     else:    # A for loop for column entries
        # for counting the values of diffrent octant i am making a new variable count
        count=0
        for xx in range(i*k,k*(i+1)):
               if xx==leng:
                    break
               
               if df['octant'][xx]==numb[j-14]:
                    count=count+1
        dict[count]=numb[j-14]
        sheet.cell(row=i+5, column=j).value =count
        sheet.cell(row=i+5, column=j).border = thin_border

   
    

    dict1 = OrderedDict(sorted(dict.items()))
    
    kk=8
    q=0
    dict2 ={}
    dict3={1:"Internal outward interaction",-1:"External outward interaction",2:"External Ejection",-2:"Internal Ejection",3:"External inward interaction",-3:"Internal inward interaction",4:"Internal sweep",-4:"External sweep"}
    for key in dict1:

          dict2[dict1[key]]=kk
          kk=kk-1
    for qq in range(1,5):
          for j in range(23,30,2):
               if(sheet.cell(row=1, column=j).value==-1*qq):
                    sheet.cell(row=i+5, column=j).value =dict2[-1*qq]
                    sheet.cell(row=i+5, column=j).border = thin_border

          for j in range(22,30,2):
               if(sheet.cell(row=1, column=j).value==qq):
                    sheet.cell(row=i+5, column=j).value =dict2[qq]
                    sheet.cell(row=i+5, column=j).border = thin_border
          if(dict2[qq]==1):
               q=qq
          if(dict2[-1*qq]==1):
               q=-1*qq 
          sheet.cell(row=i+5, column=30).value =q
          sheet.cell(row=i+5, column=30).border = thin_border
    
    sheet.cell(row=i+5, column=31).value =dict3[q]
    dict5[dict3[q]]+=1
    sheet.cell(row=i+5, column=31).border = thin_border




RR=R+8
colnew=14
print(dict5)

sheet.cell(row=RR, column=14).value ="Octant ID"
sheet.cell(row=RR, column=14).border = thin_border

sheet.cell(row=RR, column=15).value ="Octant Name "
sheet.cell(row=RR, column=15).border = thin_border

sheet.cell(row=RR, column=16).value ="Count of Rank 1 Mod Values"
sheet.cell(row=RR, column=16).border = thin_border
RR+=1
for i in range(1,5):

     sheet.cell(row=RR, column=14).value =i
     sheet.cell(row=RR, column=14).border = thin_border

     sheet.cell(row=RR, column=15).value =dict3[i]
     sheet.cell(row=RR, column=15).border = thin_border

     sheet.cell(row=RR, column=16).value =dict5[dict3[i]]
     sheet.cell(row=RR, column=16).border = thin_border
     RR+=1
     sheet.cell(row=RR, column=14).value =-1*i
     sheet.cell(row=RR, column=14).border = thin_border

     sheet.cell(row=RR, column=15).value =dict3[-1*i]
     sheet.cell(row=RR, column=15).border = thin_border

     sheet.cell(row=RR, column=16).value =dict5[dict3[-1*i]]
     sheet.cell(row=RR, column=16).border = thin_border

     RR+=1

     

print(dict2)
 






# a.append(df['octant'].value_counts()['1'])
# a.append(df['octant'].value_counts()['-1'])
# a.append(df['octant'].value_counts()['2'])
# a.append(df['octant'].value_counts()['-2'])
# a.append(df['octant'].value_counts()['3'])
# a.append(df['octant'].value_counts()['-3'])
# a.append(df['octant'].value_counts()['4'])
# a.append(df['octant'].value_counts()['-4'])


# for i in range(2,10):
     
#      for j in range(13,16):
#           # for the +1,+2,+3,+4
#           if j==13 and i%2==0:
#                sheet.cell(row=i, column=j).value =countp
#                sheet.cell(row=i, column=j).border = thin_border
             
#                maxo=0
#                count=0
#                t=0
#                for k in df['octant']:
#                     if k==countp:
#                          t=t+1
#                     else:
#                          t=0
#                     if(maxo<t):
#                          maxo=t
                    
                    
#                for k in df['octant']:
#                     if k==countp:
#                          t=t+1
#                     else:
#                          t=0
#                     if(maxo==t):
#                          count=count+1
                    
#                # adding the longest subsequece in the matrix
#                sheet.cell(row=i, column=j+1).value =maxo
#                # adding the border to it
#                sheet.cell(row=i, column=j+1).border = thin_border
#                # no times the longest susequence appears in the octant matrix
#                sheet.cell(row=i, column=j+2).value =count
#                # adding the border to it
#                sheet.cell(row=i, column=j+2).border = thin_border

#                countp=countp+1
    
#            # for the -1,-2,-3,-4  
#           if j==13 and i%2!=0:
#                sheet.cell(row=i, column=j).value =countn
#                sheet.cell(row=i, column=j).border = thin_border
               
             
#                maxo=0
#                count=0
#                t=0
#                for k in df['octant']:
#                     if k==countn:
#                          t=t+1
#                     else:
#                          t=0
#                     if(maxo<t):
#                          maxo=t
                    
                    
#                for k in df['octant']:
#                     if k==countn:
#                          t=t+1
#                     else:
#                          t=0
#                     if(maxo==t):
#                          count=count+1
                    
#                # adding the longest subsequece in the matrix
#                sheet.cell(row=i, column=j+1).value =maxo
#                # adding the border to it
#                sheet.cell(row=i, column=j+1).border = thin_border
#                 # no times the longest susequence appears in the octant matrix
#                sheet.cell(row=i, column=j+2).value =count
#                # adding the border to it
#                sheet.cell(row=i, column=j+2).border = thin_border
               
#                countn=countn-1  

# saving the final output file
book.save("output_octant_longest_subsequence.xlsx")