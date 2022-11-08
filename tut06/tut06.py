# importing required libraries
from operator import concat
import pandas as pd
import itertools
import csv
import numpy as np
import os
from datetime import datetime
from csv import writer
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
import email, smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from IPython.display import display, HTML
import datetime  
from datetime import date 
import smtplib

# reading the input csv file
df=pd.read_csv('input_attendance.csv')

dt=pd.read_csv('input_registered_students.csv')

k=df['Attendance'].isnull().sum()  
print(k)
                                                                                                            
# some the value are missing so i have to remove those values from my data set
df['Attendance'].replace('', np.nan, inplace=True)
print(df)
df.dropna(subset=['Attendance'], inplace=True)
print(df)

# after removing the Null values we again reading the data set with no null values
df.to_csv('input_attendance.csv', mode='w', header=True,index=False)
df=pd.read_csv('input_attendance.csv')

# dropping null value columns to avoid errors
df.dropna(inplace = True)
#  in the below code we extracting the date and time from timestamp
# new data frame with split value columns
new = df["Timestamp"].str.split(" ", n = 1, expand = True)
# spliting the attendance column into roll no and name column
rollna = df["Attendance"].str.split(" ", n = 1, expand = True)
df["Roll"]=rollna[0]
df["Name"]=rollna[1]
# making separate first name column for roll no from new data frame
df["Date"]= new[0]
 
# making separate last name column for name from new data frame
df["time"]= new[1]

# extracting the hours from time column
hour = df["time"].str.split(":", n = 1, expand = True)
df['hour']=hour[0]
 
# Dropping old attendance and time columns

df.drop(columns =["time"], inplace = True)
df.drop(columns =["Attendance"], inplace = True)

#  in the name and roll no column may values are in lower and uper case so we make all the values in upper case

df['Name'] = df['Name'].str.upper()
df['Roll'] = df["Roll"].str.upper()
a=[]

# in the below code we are trying to know the day wether  it is monday or thrusday or any other day
print(df["Date"])
for x in df["Date"]:
    date=x

    day, month, year = date.split('-')     
    day_name = datetime.date(int(year), int(month), int(day)) 
    a.append(day_name.strftime("%A"))
df["days"]=a

leng=df["Date"].count()




# importing packages

  
# range of timestamps
k = pd.date_range(start=df["Timestamp"].iloc[0],
                                end=df["Timestamp"].iloc[-1])

dato=[]
ak=[]
for i in k:
    
    az=str(i).split(" ")
    dato.append(i.strftime('%d-%m-20%y'))
    
    d = pd.Timestamp(az[0])
    ak.append(d.day_name())
    


for i,j in zip(dato,ak):
    print(i,j)

df.to_excel('ouyputt.xlsx',index=False)



dataf=[]
ann=[]




# Pass the list as an argument into
datee=[]
total_lec=0
datee.append("Roll NO")
datee.append("Name")
rr=" "
for i,j in zip(dato,ak):
    if(j=='Thursday'or j=='Monday'):
        if i!=rr:
            rr=i
            total_lec+=1
            datee.append(i)
            ann.append(i)


datee.append("Actual Lecture Taken") 
datee.append("Total Real")
datee.append("% Attendance") 


print(datee)
with open('output\Attendance_report_consolidated.csv','w') as f_object:
 
    # Pass this file object to csv.writer()
    # and get a writer object
    writer_object = writer(f_object)
 
    # Pass the list as an argument into
    # the writerow()
    writer_object.writerow(datee)
    


 
    # Pass this file object to csv.writer()
    # and get a writer object
    
 
    # Pass the list as an argument into
    # the writerow()
    for k,l in zip(dt['Roll No'],dt['Name']):
            list=[]
            list.append(k)
            list.append(l)
            realtek=0
            for i in ann:
                flag=0
                for j,e,o in zip(df["Date"],df["hour"],df["Roll"]):

                    if i==j and e=="14" and k==o:
                        list.append("P")
                        realtek+=1
                        flag=1
                        break
                        

                if flag==0:
                    list.append("A")
  
            list.append(total_lec) 
            list.append(realtek)
            list.append(round((((realtek)/total_lec)*100),2))            
            writer_object.writerow(list)
            

    f_object.close()

dk=pd.read_csv('output\\attendance_report_consolidated.csv')
dk['Name'].replace('', np.nan, inplace=True)
print(dk)
dk.dropna(subset=['Name'], inplace=True)
dk.to_excel('output\\attendance_report_consolidated.xlsx',index=False)

os.remove('output\\attendance_report_consolidated.csv')


# inmporting nessesry libreries
from ast import Try
from operator import concat
import pandas as pd
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook
#load excel file




# import xlsxwriter module
import xlsxwriter
 
# Workbook() takes one, non-optional, argument
# which is the filename that we want to create.
for i in dt["Roll No"]:
            workbook = xlsxwriter.Workbook("output\\"+i+".xlsx")
            
            # The workbook object is then used to add new
            # worksheet via the add_worksheet() method.
            worksheet = workbook.add_worksheet()
            
            # Use the worksheet object to write
            # data via the write() method.
            worksheet.write('A1', 'Hello..')
            worksheet.write('B1', 'Geeks')
            worksheet.write('C1', 'For')
            worksheet.write('D1', 'Geeks')
            
            # Finally, close the Excel file
            # via the close() method.
            workbook.close()
 
#definging the property of border
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
for k,l in zip(dt['Roll No'],dt['Name']):
            
                er=k 
                book = load_workbook(filename="output\\"+er+".xlsx")
                sheet = book.active
                sheet.cell(row=1, column=1).value ="Date"
                sheet.cell(row=1, column=2).value ="Roll No"
                sheet.cell(row=1, column=3).value ="Name"
                sheet.cell(row=1, column=4).value ="Total Attendance Count"
                sheet.cell(row=1, column=5).value ="Real"
                sheet.cell(row=1, column=6).value ="Duplicate"
                sheet.cell(row=1, column=7).value ="Invalid"
                sheet.cell(row=1, column=8).value ="Absent"
                sheet.cell(row=2, column=2).value =k
                sheet.cell(row=2, column=3).value =l
                aw=3
                
                for i in ann:
                    
                    
                    sheet.cell(row=aw, column=1).value =i
                    flag=0
                    count=0
                    invalid=0
                    for j,e,o in zip(df["Date"],df["hour"],df["Roll"]):

                        if i==j and e=="14" and k==o:
                            count+=1
                            
                            
                        elif i==j and k==o: 
                            invalid+=1
                    sheet.cell(row=aw, column=4).value =count+invalid
                    if count>0:
                        sheet.cell(row=aw, column=5).value =1
                        sheet.cell(row=aw, column=6).value =count-1
                        sheet.cell(row=aw, column=8).value =0
                    else:
                        sheet.cell(row=aw, column=5).value =0
                        sheet.cell(row=aw, column=6).value =0
                        sheet.cell(row=aw, column=8).value =1
                    
                    
                    sheet.cell(row=aw, column=7).value =invalid      
                    aw+=1
                    
                    book.save("output\\"+er+".xlsx")   
            
            
               
                
            
            #open workbook
            





            

# ------------------------------------------------taking input file---------------------------------------------------------------
# appling try and except for  checking the file 

















     
    
# the writerow()






# def send_email():                                                                           # Function to send email to cs3842022@gmail.com
#         try:
#             subject = "Consolidated Attendace Report"                                           
#             sender_email = input("Enter sender email : ")                                       # Sender e-mail
#             receiver_email = "cs3842022@gmail.com"                                        # Receiver e-mail => cs3842022@gmail.com
#             password = input("Type your password and press enter:")                             # Password of sender e-mail
#             body = "The report is attached "
#             # Create a multipart message and set headers
#             message = MIMEMultipart()                                                       
#             message["From"] = sender_email
#             message["To"] = receiver_email
#             message["Subject"] = subject
#             message["Bcc"] = receiver_email  # Recommended for mass emails

#             # Add body to email
#             message.attach(MIMEText(body, "plain"))

#             filename = "output\Attendance_report_consolidated.csv"  # In same directory as script

#             # Open csv file in binary mode
#             with open(filename, "rb") as attachment:
#                 # Add file as application/octet-stream
#                 # Email client can usually download this automatically as attachment
#                 part = MIMEBase("application", "octet-stream")
#                 part.set_payload(attachment.read())

#             # Encode file in ASCII characters to send by email    
#             encoders.encode_base64(part)

#             # Add header as key/value pair to attachment part
#             part.add_header(
#                 "Content-Disposition",
#                 f"attachment; filename= {filename}",
#             )
#             # Add attachment to message and convert message to string
#             message.attach(part)
#             text = message.as_string()

#             # Log in to server using secure context and send email
#             context = ssl.create_default_context()
#             with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
#                 server.login(sender_email, password)
#                 server.sendmail(sender_email, receiver_email, text)
#                 print("succsessfully sended the file :)")
#         except:
           
#             print("Error in sending the file.")
# send_email()
# # tbvvrtbtmqivlliq