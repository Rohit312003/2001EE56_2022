

# inmporting nessesry libreries
from ast import Try
from operator import concat
import pandas as pd
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook
#load excel file
try:
     book = load_workbook(filename="input_octant_longest_subsequence_with_range.xlsx")
except:
     print("unable to find the input file")
     exit()
 
#open workbook
sheet = book.active
 
#definging the property of border
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# ------------------------------------------------taking input file---------------------------------------------------------------
# appling try and except for  checking the file 
try:
                                                                      #reading the input file and storing its value in the df 
     df = pd.read_excel('input_octant_longest_subsequence_with_range.xlsx')
except:
     print("unable to find the input file")
     exit()

print("Column headings:")
print(df)
# ----------------------------------------------calculating the value of mean of column U,V and W-----------------------------------
try:
     meanU=df['U'].mean()
     meanV=df['V'].mean()
     meanW=df['W'].mean()
except:
     print("unable to find the mean")

sheet = book.active


meanU=df['U'].mean()
meanV=df['V'].mean()
meanW=df['W'].mean()
# adding the mean value to the new column
sheet.cell(row=1, column=5).value ="Uavg"
sheet.cell(row=2, column=5).value =meanU

sheet.cell(row=1, column=6).value ="Vavg"
sheet.cell(row=2, column=6).value =meanV

sheet.cell(row=1, column=7).value ="Wavg"
sheet.cell(row=2, column=7).value =meanW 
# saving the output to the output file
book.save("output_octant_longest_subsequence_with_range.xlsx")
# reading the same saved output value to new again and store it into df
df = pd.read_excel('output_octant_longest_subsequence_with_range.xlsx')
# ----------------------------------------------  calculating error in U,V and W----------------------------------
df['U-Uavg']=df['U']-meanU
df['V-Vavg']=df['V']-meanV
df['W-Wavg']=df['W']-meanW
# ----------------------------------------------below code for desiding the octant----------------------------------------------------------
                                                                  # creating a list for octant values
octant=[]                                                        
 #  calculating the Octant value .
try:
     for (l, m, q) in zip(df['U-Uavg'],df['V-Vavg'],df['W-Wavg']):     #   appanding the octannt values to the list
         if l>0 and m>0 and q<0:
             octant.append(-1)
         elif l>0 and m>0 and q>0:
              octant.append(1)
         elif l<0 and m>0 and q<0:
              octant.append(-2)
         elif l<0 and m>0 and q>0:
              octant.append(2)
         elif l<0 and m<0 and q<0:
              octant.append(-3)
         elif l<0 and m<0 and q>0:
           octant.append(3)
         elif l>=0 and m<=0 and q<0:
            octant.append(-4)
         elif l>0 and m<0 and q>0:
             octant.append(4) 
     df['octant'] = octant
except:
     print("somting wrong with the columns U-Uavg,V-Vavg,W-Wavg ")

# saving the file in out put file
with pd.ExcelWriter('output_octant_longest_subsequence_with_range.xlsx') as writer:
    df.to_excel(writer, sheet_name='Sheet_name_1',index=False)
print("sucssesfully done!")
# reading the same file again  in workbook

book = load_workbook(filename="output_octant_longest_subsequence_with_range.xlsx")
 
#open workbook
sheet = book.active

# making our desire matrix having desired heading
sheet.cell(row=2, column=13).value ="Count"
sheet.cell(row=2, column=13).border = thin_border

sheet.cell(row=2, column=14).value ="Longest Subsquence Length"
sheet.cell(row=2, column=14).border = thin_border

sheet.cell(row=2, column=15).value ="Count"
sheet.cell(row=2, column=15).border = thin_border
# making our desire matrix having desired heading
sheet.cell(row=2, column=17).value ="Count"
sheet.cell(row=2, column=17).border = thin_border

sheet.cell(row=2, column=18).value ="Longest Subsquence Length"
sheet.cell(row=2, column=18).border = thin_border

sheet.cell(row=2, column=19).value ="Count"
sheet.cell(row=2, column=19).border = thin_border
# ----------------------below code is for calculating the longest subsequence ans its count---------------------
countp=1
countn=-1

for i in range(3,11):
     
     for j in range(13,16):
          # for the +1,+2,+3,+4
          if j==13 and i%2==0:
               sheet.cell(row=i, column=j).value =countp
               sheet.cell(row=i, column=j).border = thin_border
             
               maxo=0
               count=0
               t=0
               for k in df['octant']:
                    if k==countp:
                         t=t+1
                    else:
                         t=0
                    if(maxo<t):
                         maxo=t
                    
                    
               for k in df['octant']:
                    if k==countp:
                         t=t+1
                    else:
                         t=0
                    if(maxo==t):
                         count=count+1
                    
               # adding the longest subsequece in the matrix
               sheet.cell(row=i, column=j+1).value =maxo
               # adding the border to it
               sheet.cell(row=i, column=j+1).border = thin_border
               # no times the longest susequence appears in the octant matrix
               sheet.cell(row=i, column=j+2).value =count
               # adding the border to it
               sheet.cell(row=i, column=j+2).border = thin_border

               countp=countp+1
    
           # for the -1,-2,-3,-4  
          if j==13 and i%2!=0:
               sheet.cell(row=i, column=j).value =countn
               sheet.cell(row=i, column=j).border = thin_border
               
             
               maxo=0
               count=0
               t=0
               for k in df['octant']:
                    if k==countn:
                         t=t+1
                    else:
                         t=0
                    if(maxo<t):
                         maxo=t
                    
                    
               for k in df['octant']:
                    if k==countn:
                         t=t+1
                    else:
                         t=0
                    if(maxo==t):
                         count=count+1
                    
               # adding the longest subsequece in the matrix
               sheet.cell(row=i, column=j+1).value =maxo
               # adding the border to it
               sheet.cell(row=i, column=j+1).border = thin_border
                # no times the longest susequence appears in the octant matrix
               sheet.cell(row=i, column=j+2).value =count
               # adding the border to it
               sheet.cell(row=i, column=j+2).border = thin_border
               
               
               countn=countn-1  
# -------------------below code for calculating the longest subsequence ans its count  with time included--------------------------------------------------------

#  making a new variable to keep track of column so that i can easly add new columns
star=3
for i in range(1,5):
     
     for j in range(17,18):
          # for the +1,+2,+3,+4
          if j==17 :

               sheet.cell(row=star, column=j).value =1*i
               sheet.cell(row=star, column=j).border = thin_border
             
               maxo=0
               count=0
               t=0
               # below code for calulating the max subsequence length
               for k in df['octant']:
                    if k==i:
                         t=t+1
                    else:
                         t=0

                    if(maxo<t):
                         maxo=t
                    
               # below code for calulating the no frequency for occurence max subsequence length   
               for k in df['octant']:
                    if k==i:
                         t=t+1
                    else:
                         t=0
                    if(maxo==t):
                         count=count+1
                    
               # adding the longest subsequece in the matrix
               sheet.cell(row=star, column=j+1).value =maxo
               # adding the border to it
               sheet.cell(row=star, column=j+1).border = thin_border
               # no times the longest susequence appears in the octant matrix
               sheet.cell(row=star, column=j+2).value =count
               # adding the border to it
               sheet.cell(row=star, column=j+2).border = thin_border
               # increasing the value of varble to keep track of no. columns
               star=star+1

               # adding the new row 
               # for time to from.
               sheet.cell(row=star, column=j).value ="Time"
               # adding the border to it
               sheet.cell(row=star, column=j).border = thin_border
               sheet.cell(row=star, column=j+1).value ="From"
               # adding the border to it
               sheet.cell(row=star, column=j+1).border = thin_border
               sheet.cell(row=star, column=j+2).value ="To"
               # adding the border to it
               sheet.cell(row=star, column=j+2).border = thin_border
                # increasing the value of varble to keep track of no. columns
               star=star+1

               # below code for finding where the max subsequnce length starts and ends.
               # below variable to keep track no itrations whilch halps to find the dirst occurence of an element ifrom the max subsequence length
               qq=0
               for k in df['octant']:
                    qq=qq+1
                    if k==i:
                         t=t+1
                    else:
                         t=0
                    if(maxo==t):
                         sheet.cell(row=star, column=j).value =" "
                         # adding the border to it
                         sheet.cell(row=star, column=j).border = thin_border
                         sheet.cell(row=star, column=j+1).value =df['Time'][qq-maxo]
                         # adding the border to it
                         sheet.cell(row=star, column=j+1).border = thin_border
                         sheet.cell(row=star, column=j+2).value =df['Time'][qq-1]
                         # adding the border to it
                         sheet.cell(row=star, column=j+2).border = thin_border

                          # increasing the value of varble to keep track of no columns
                         star=star+1
                         
                    
          # same code as above only for negatve of i or(if above is for 1 than this below code is for -1 and so on).
          if j==17 :
               i=-1*i
               sheet.cell(row=star, column=j).value =1*i
               sheet.cell(row=star, column=j).border = thin_border
             
               maxo=0
               count=0
               t=0
               for k in df['octant']:
                    if k==i:
                         t=t+1
                    else:
                         t=0

                    if(maxo<t):
                         maxo=t
                    
                    
               for k in df['octant']:
                    if k==i:
                         t=t+1
                    else:
                         t=0
                    if(maxo==t):
                         count=count+1
                    
               # adding the longest subsequece in the matrix
               sheet.cell(row=star, column=j+1).value =maxo
               # adding the border to it
               sheet.cell(row=star, column=j+1).border = thin_border
               # no times the longest susequence appears in the octant matrix
               sheet.cell(row=star, column=j+2).value =count
               # adding the border to it
               sheet.cell(row=star, column=j+2).border = thin_border
                 # increasing the value of varble to keep track of no columns
               star=star+1

               # adding the new row 
               # for time to from.
               sheet.cell(row=star, column=j).value ="Time"
               # adding the border to it
               sheet.cell(row=star, column=j).border = thin_border
               sheet.cell(row=star, column=j+1).value ="From"
               # adding the border to it
               sheet.cell(row=star, column=j+1).border = thin_border
               sheet.cell(row=star, column=j+2).value ="To"
               # adding the border to it
               sheet.cell(row=star, column=j+2).border = thin_border
               # increasing the value of varble to keep track of no columns
               star=star+1

               # below code for finding where the max subsequnce length starts and ends.
               # below variable to keep track no itrations whilch halps to find the dirst occurence of an element ifrom the max subsequence length
               qq=0
               for k in df['octant']:
                    qq=qq+1
                    if k==i:
                         t=t+1
                    else:
                         t=0
                    if(maxo==t):
                         sheet.cell(row=star, column=j).value =" "
                         # adding the border to it
                         sheet.cell(row=star, column=j).border = thin_border
                         sheet.cell(row=star, column=j+1).value =df['Time'][qq-maxo]
                         # adding the border to it
                         sheet.cell(row=star, column=j+1).border = thin_border
                         sheet.cell(row=star, column=j+2).value =df['Time'][qq-1]
                         # adding the border to it
                         sheet.cell(row=star, column=j+2).border = thin_border
                         star=star+1
                         
                         
     i=-1*i 
              
# saving the final output file
book.save("output_octant_longest_subsequence_with_range.xlsx")