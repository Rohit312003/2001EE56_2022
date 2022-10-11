

# inmporting nessesry libreries
from ast import Try
from operator import concat
import pandas as pd
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook


 
#load excel file
try:
     book = load_workbook(filename="input_octant_longest_subsequence.xlsx")
except:
     print("unable to find the input file")
     exit()
 
#open workbook
sheet = book.active
 
#definging the property of border
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# ------------------------------------------------taking input file---------------------------------------------------------------
# appling try and except for  checking the file 
try:
                                                                      #reading the input file and storing its value in the df 
     df = pd.read_excel('input_octant_longest_subsequence.xlsx')
except:
     print("unable to find the input file")
     exit()

print("Column headings:")
print(df)
# ----------------------------------------------calculating the value of mean of column U,V and W-----------------------------------
try:
     meanU=df['U'].mean()
     meanV=df['V'].mean()
     meanW=df['W'].mean()
except:
     print("unable to find the mean")

sheet = book.active


meanU=df['U'].mean()
meanV=df['V'].mean()
meanW=df['W'].mean()
# adding the mean value to the new column
sheet.cell(row=1, column=5).value ="Uavg"
sheet.cell(row=2, column=5).value =meanU

sheet.cell(row=1, column=6).value ="Vavg"
sheet.cell(row=2, column=6).value =meanV

sheet.cell(row=1, column=7).value ="Wavg"
sheet.cell(row=2, column=7).value =meanW 


# saving the output to the output file
book.save("output_octant_longest_subsequence.xlsx")
# reading the same saved output value to new again and store it into df
df = pd.read_excel('output_octant_longest_subsequence.xlsx')
# ----------------------------------------------  calculating error in U,V and W----------------------------------
df['U-Uavg']=df['U']-meanU
df['V-Vavg']=df['V']-meanV
df['W-Wavg']=df['W']-meanW
# ----------------------------------------------below code for desiding the octant----------------------------------------------------------
                                                                  # creating a list for octant values
octant=[]                                                        
 #  calculating the Octant value .
try:
     for (l, m, q) in zip(df['U-Uavg'],df['V-Vavg'],df['W-Wavg']):     #   appanding the octannt values to the list
         if l>0 and m>0 and q<0:
             octant.append(-1)
         elif l>0 and m>0 and q>0:
              octant.append(1)
         elif l<0 and m>0 and q<0:
              octant.append(-2)
         elif l<0 and m>0 and q>0:
              octant.append(2)
         elif l<0 and m<0 and q<0:
              octant.append(-3)
         elif l<0 and m<0 and q>0:
           octant.append(3)
         elif l>=0 and m<=0 and q<0:
            octant.append(-4)
         elif l>0 and m<0 and q>0:
             octant.append(4) 
     df['octant'] = octant
except:
     print("somting wrong with the columns U-Uavg,V-Vavg,W-Wavg ")

# saving the file in out put file
with pd.ExcelWriter('output_octant_longest_subsequence.xlsx') as writer:
    df.to_excel(writer, sheet_name='Sheet_name_1',index=False)
print("sucssesfully done!")
# reading the same file again  in workbook

book = load_workbook(filename="output_octant_longest_subsequence.xlsx")
 
#open workbook
sheet = book.active

# making our desire matrix having desired heading
sheet.cell(row=1, column=13).value ="Count"
sheet.cell(row=1, column=13).border = thin_border

sheet.cell(row=1, column=14).value ="Longest Subsquence Length"
sheet.cell(row=1, column=14).border = thin_border

sheet.cell(row=1, column=15).value ="Count"
sheet.cell(row=1, column=15).border = thin_border


# below code is for calculating the longest subsequence ans its count
countp=1
countn=-1

for i in range(2,10):
     
     for j in range(13,16):
          # for the +1,+2,+3,+4
          if j==13 and i%2==0:
               sheet.cell(row=i, column=j).value =countp
               sheet.cell(row=i, column=j).border = thin_border
             
               maxo=0
               count=0
               t=0
               for k in df['octant']:
                    if k==countp:
                         t=t+1
                    else:
                         t=0
                    if(maxo<t):
                         maxo=t
                    
                    
               for k in df['octant']:
                    if k==countp:
                         t=t+1
                    else:
                         t=0
                    if(maxo==t):
                         count=count+1
                    
               # adding the longest subsequece in the matrix
               sheet.cell(row=i, column=j+1).value =maxo
               # adding the border to it
               sheet.cell(row=i, column=j+1).border = thin_border
               # no times the longest susequence appears in the octant matrix
               sheet.cell(row=i, column=j+2).value =count
               # adding the border to it
               sheet.cell(row=i, column=j+2).border = thin_border

               countp=countp+1
    
           # for the -1,-2,-3,-4  
          if j==13 and i%2!=0:
               sheet.cell(row=i, column=j).value =countn
               sheet.cell(row=i, column=j).border = thin_border
               
             
               maxo=0
               count=0
               t=0
               for k in df['octant']:
                    if k==countn:
                         t=t+1
                    else:
                         t=0
                    if(maxo<t):
                         maxo=t
                    
                    
               for k in df['octant']:
                    if k==countn:
                         t=t+1
                    else:
                         t=0
                    if(maxo==t):
                         count=count+1
                    
               # adding the longest subsequece in the matrix
               sheet.cell(row=i, column=j+1).value =maxo
               # adding the border to it
               sheet.cell(row=i, column=j+1).border = thin_border
                # no times the longest susequence appears in the octant matrix
               sheet.cell(row=i, column=j+2).value =count
               # adding the border to it
               sheet.cell(row=i, column=j+2).border = thin_border
               
               countn=countn-1  

# saving the final output file
book.save("output_octant_longest_subsequence.xlsx")