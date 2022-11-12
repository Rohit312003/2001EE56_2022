# inmporting nessesry libreries
import openpyxl
from openpyxl.styles import PatternFill
from collections import OrderedDict
from ast import Try
from operator import concat
import numpy as np
import pandas as pd
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook
import glob
import os
# getting excel files from Directory Desktop
path = "input\\"
# read all the files with extension .xlsx i.e. excel 
filenames = glob.glob(path + "\*.xlsx")
print('File names:', filenames)
# for loop to iterate all excel files 
from datetime import datetime
start_time = datetime.now()
for file in filenames:
	print(file)
	#load excel file
	try:
		book = load_workbook(filename=file)
	except:
		print("unable to find the input file")
		exit()
	#open workbook
	sheet = book.active
	#definging the property of border
	thin_border = Border(left=Side(style='thin'), 
						right=Side(style='thin'), 
						top=Side(style='thin'), 
						bottom=Side(style='thin'))

	# -------------------------------------------------------------taking input file---------------------------------------------------------------
	# appling try and except for  checking the file 
	try:
																		#reading the input file and storing its value in the df 
		df = pd.read_excel(file)
	except:
		print("unable to find the input file")
		exit()

	print("Column headings:")
	print(df)
	# ------------------------------------------------------calculating the value of mean of column U,V and W-----------------------------------
	try:
		meanU=df['U'].mean()
		meanV=df['V'].mean()
		meanW=df['W'].mean()
	except:
		print("unable to find the mean")
	sheet = book.active
	meanU=df['U'].mean()
	meanV=df['V'].mean()
	meanW=df['W'].mean()
	# adding the mean value to the new column
	sheet.cell(row=1, column=5).value ="Uavg"
	sheet.cell(row=2, column=5).value =round(meanU,3)
	sheet.cell(row=1, column=6).value ="Vavg"
	sheet.cell(row=2, column=6).value =round(meanV,3)
	sheet.cell(row=1, column=7).value ="Wavg"
	sheet.cell(row=2, column=7).value =round(meanW,3)
	# ----------------------------------------------  calculating error in U,V and W----------------------------------
	df['U-Uavg']=df['U']-meanU
	df['V-Vavg']=df['V']-meanV
	df['W-Wavg']=df['W']-meanW
	# -------------------------------------------------------------------------------------------------------															# creating a list for octant values
	octant=[]                                                        
	#  calculating the value of U-Uavg, V-Uavg and W-Wavg
	sheet.cell(row=1, column=8).value ="U-Uavg"
	kkk=2
	for i in df['U-Uavg']:
		sheet.cell(row=kkk, column=8).value =i
		kkk+=1
	sheet.cell(row=1, column=9).value ="V-Vavg"
	kkk=2
	for i in df['V-Vavg']:
			sheet.cell(row=kkk, column=9).value =i
			kkk+=1
	sheet.cell(row=1, column=10).value ="W-Wavg"
	kkk=2
	for i in df['W-Wavg']:
			sheet.cell(row=kkk, column=10).value =i
			kkk+=1
	# desiding the octant value
	try:
		for (l, m, q) in zip(df['U-Uavg'],df['V-Vavg'],df['W-Wavg']):
				#   appanding the octannt values to the list
			try:
				if l>0 and m>0 and q<0:
						octant.append(-1)
				elif l>0 and m>0 and q>0:
						octant.append(1)
				elif l<0 and m>0 and q<0:
						octant.append(-2)
				elif l<0 and m>0 and q>0:
						octant.append(2)
				elif l<0 and m<0 and q<0:
						octant.append(-3)
				elif l<0 and m<0 and q>0:
						octant.append(3)
				elif l>=0 and m<=0 and q<0:
						octant.append(-4)
				elif l>0 and m<0 and q>0:
						octant.append(4) 
			except:
				print("error in appending the value")
		df['octant'] = octant
	except:
		print("somting wrong with the columns U-Uavg,V-Vavg,W-Wavg ")
	# making the desire heading
	sheet.cell(row=4, column=12).value ="User input"
	sheet.cell(row=4, column=12).border = thin_border
	# mod value
	mod=5000
	sheet.cell(row=4, column=13).value ="mod "+str(mod)
	sheet.cell(row=4, column=13).border = thin_border
	# making our desire desired heading
	sheet.cell(row=1, column=13).value ="Overall Octant Count"
	sheet.cell(row=1, column=13).border = thin_border
	sheet.cell(row=2, column=13).value ="Octant ID"
	sheet.cell(row=2, column=13).border = thin_border
	sheet.cell(row=2, column=14).value =1
	sheet.cell(row=2, column=14).border = thin_border
	sheet.cell(row=2, column=15).value =-1
	sheet.cell(row=2, column=15).border = thin_border
	sheet.cell(row=2, column=16).value =2
	sheet.cell(row=2, column=16).border = thin_border
	sheet.cell(row=2, column=17).value =-2
	sheet.cell(row=2, column=17).border = thin_border
	sheet.cell(row=2, column=18).value =3
	sheet.cell(row=2, column=18).border = thin_border
	sheet.cell(row=2, column=19).value =-3
	sheet.cell(row=2, column=19).border = thin_border
	sheet.cell(row=2, column=20).value =4
	sheet.cell(row=2, column=20).border = thin_border
	sheet.cell(row=2, column=21).value =-4
	sheet.cell(row=2, column=21).border = thin_border
	sheet.cell(row=1, column=22).value =1
	sheet.cell(row=1, column=22).border = thin_border
	sheet.cell(row=2, column=22).value ="Rank 1"
	sheet.cell(row=2, column=22).border = thin_border
	sheet.cell(row=1, column=23).value =-1
	sheet.cell(row=1, column=23).border = thin_border
	sheet.cell(row=2, column=23).value ="Rank -1"
	sheet.cell(row=2, column=23).border = thin_border
	sheet.cell(row=1, column=24).value =2
	sheet.cell(row=1, column=24).border = thin_border
	sheet.cell(row=2, column=24).value ="Rank 2"
	sheet.cell(row=2, column=24).border = thin_border
	sheet.cell(row=1, column=25).value =-2
	sheet.cell(row=1, column=25).border = thin_border
	sheet.cell(row=2, column=25).value ="Rank -2"
	sheet.cell(row=2, column=25).border = thin_border
	sheet.cell(row=1, column=26).value =3
	sheet.cell(row=1, column=26).border = thin_border
	sheet.cell(row=2, column=26).value ="Rank 3"
	sheet.cell(row=2, column=26).border = thin_border
	sheet.cell(row=1, column=27).value =-3
	sheet.cell(row=1, column=27).border = thin_border
	sheet.cell(row=2, column=27).value ="Rank -3"
	sheet.cell(row=2, column=27).border = thin_border
	sheet.cell(row=1, column=28).value =4
	sheet.cell(row=1, column=28).border = thin_border
	sheet.cell(row=2, column=28).value ="Rank 4"
	sheet.cell(row=2, column=28).border = thin_border
	sheet.cell(row=1, column=29).value =-4
	sheet.cell(row=1, column=29).border = thin_border
	sheet.cell(row=2, column=29).value ="Rank -4"
	sheet.cell(row=2, column=29).border = thin_border
	sheet.cell(row=2, column=30).value ="Rank1 Octant ID"
	sheet.cell(row=2, column=30).border = thin_border
	sheet.cell(row=2, column=31).value ="Rank1 Octant Name"
	sheet.cell(row=2, column=31).border = thin_border
	sheet.cell(row=3, column=13).value ="Overall Count"
	sheet.cell(row=3, column=13).border = thin_border
	# below code is for calculating the longest subsequence ans its count
	dict = {}
	for k in range(1,5):
		maxo=0
		maxi=0
		for i in range(14,22,2):
		
			if(sheet.cell(row=2, column=i).value==k):
				sheet.cell(row=3, column=i).value =df['octant'].value_counts()[k]
				sheet.cell(row=3, column=i).border = thin_border
				dict[df['octant'].value_counts()[k]]=k
		for i in range(15,22,2):
		
			if(sheet.cell(row=2, column=i).value==-1*k):
				sheet.cell(row=3, column=i).value =df['octant'].value_counts()[-1*k]
				sheet.cell(row=3, column=i).border = thin_border
				dict[df['octant'].value_counts()[-1*k]]=-1*k
	# Creates a sorted dictionary (sorted by key)
	# sorting the disctionery
	dict1 = OrderedDict(sorted(dict.items()))
	k=8
	q=0
	dict2 ={}

	for key in dict1:
		dict2[dict1[key]]=k
		k=k-1
	# below code for octant count(overall)
	for k in range(1,5):
		print(k)
		for i in range(23,30,2):
			if(sheet.cell(row=1, column=i).value==-1*k):
				sheet.cell(row=3, column=i).value =dict2[-1*k]
				if(dict2[-1*k]==1):
					# below code for coloring of element with specific  row and column
					sheet.cell(row=3,column=i).fill = PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid") 
				sheet.cell(row=3, column=i).border = thin_border
		for i in range(22,30,2):
			if(sheet.cell(row=1, column=i).value==k):
				sheet.cell(row=3, column=i).value =dict2[k]
				if(dict2[k]==1):
					sheet.cell(row=3,column=i).fill = PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")
				sheet.cell(row=3, column=i).border = thin_border
		if(dict2[k]==1):
			q=k 
		elif(dict2[-1*k]==1):
			q=-1*k 
	dict0={1:"Internal outward interaction",-1:"External outward interaction",2:"External Ejection",-2:"Internal Ejection",3:"External inward interaction",-3:"Internal inward interaction",4:"Internal sweep",-4:"External sweep"}
	sheet.cell(row=3, column=30).value =q
	sheet.cell(row=3, column=30).border = thin_border
	sheet.cell(row=3, column=31).value =dict0[q]
	sheet.cell(row=3, column=31).border = thin_border	
	try:
		leng=df['octant'].count()
	except:
		print("errpr in leng")
	
	# putting the value of mod in k ki s equal to 5000
	k=mod
	z=int(leng/k)+1
	# no rows in output
	R = int(z)
	# column where my loop ends
	C = int(22)
	# creating the list for octant value
	numb=[1,-1,2,-2,3,-3,4,-4]
	# making a desire dectionary calulation of the output below box
	dict5={"Internal outward interaction":0,"External outward interaction":0,"External Ejection":0,"Internal Ejection":0,"External inward interaction":0,"Internal inward interaction":0,"Internal sweep":0,"External sweep":0}


	for i in range(R):
				
	
		dict = {}
		for j in range(13,C): 
			if j==13:
				# condition for last row of output to cheak the  length of input so it stop when input length reached
				if i==R-1:
					u=str(i*k)+"-"+str(leng) 
					sheet.cell(row=i+5, column=j).value =u
					sheet.cell(row=i+5, column=j).border = thin_border
				else:         
					u=str(i*k)+"-"+str(k*(i+1)-1) 
					sheet.cell(row=i+5, column=j).value =u
					sheet.cell(row=i+5, column=j).border = thin_border
			else:    
				# for counting the values of diffrent octant i am making a new variable count
				count=0
				for xx in range(i*k,k*(i+1)):
					if xx==leng:
							break
					
					if df['octant'][xx]==numb[j-14]:
							count=count+1
			# below code for storing the value of noumber of count of certain octant in a dectionary count:octant manner
				dict[count]=numb[j-14]
				sheet.cell(row=i+5, column=j).value =count
				sheet.cell(row=i+5, column=j).border = thin_border
	#     sorting the dectionary by its value order
		dict1 = OrderedDict(sorted(dict.items()))
		# using this vable to find the rank in a order
		kk=8
	#     below variable is for Octant name 
		q=0
	#     making new dectionary for storing the rank in desired order
		dict3={1:"Internal outward interaction",-1:"External outward interaction",2:"External Ejection",-2:"Internal Ejection",3:"External inward interaction",-3:"Internal inward interaction",4:"Internal sweep",-4:"External sweep"}
		for key in dict1:
		#     storing the octant value with  value equal its rank decreasing order
			dict2[dict1[key]]=kk
			kk=kk-1
		# below code for the matrix adter input value    
		for qq in range(1,5):
		# positive 1,2,3,4
			for j in range(23,30,2):
				if(sheet.cell(row=1, column=j).value==-1*qq):
						sheet.cell(row=i+5, column=j).value =dict2[-1*qq]
						if(dict2[-1*qq]==1):
							sheet.cell(row=i+5,column=j).fill = PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")
						sheet.cell(row=i+5, column=j).border = thin_border
	#    for negative -1,-2,-3,-4
			for j in range(22,30,2):
				if(sheet.cell(row=1, column=j).value==qq):
						sheet.cell(row=i+5, column=j).value =dict2[qq]
						if(dict2[qq]==1):
							sheet.cell(row=i+5,column=j).fill = PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")
						
						sheet.cell(row=i+5, column=j).border = thin_border
			# below for rank 1 octant value
			if(dict2[qq]==1):
				q=qq
				print(q)
			elif(dict2[-1*qq]==1):
				q=-1*qq
				print(q)
			else:
				q=1
			# wite in a cell
			sheet.cell(row=i+5, column=30).value =q
			sheet.cell(row=i+5, column=30).border = thin_border
		sheet.cell(row=i+5, column=31).value =dict3[q]
		dict5[dict3[q]]+=1
		sheet.cell(row=i+5, column=31).border = thin_border
	# below variable treck record of no colmns for helping the postion for below matrix
	RR=R+8
	# below matrix where its colmns starts
	colnew=14
	# heading of below matrix
	sheet.cell(row=RR, column=14).value ="Octant ID"
	sheet.cell(row=RR, column=14).border = thin_border
	sheet.cell(row=RR, column=15).value ="Octant Name "
	sheet.cell(row=RR, column=15).border = thin_border
	sheet.cell(row=RR, column=16).value ="Count of Rank 1 Mod Values"
	sheet.cell(row=RR, column=16).border = thin_border
	# keep tracing no rows
	RR+=1
	# below matrix vlue calculating
	for i in range(1,5):
	# for positve value i
		sheet.cell(row=RR, column=14).value =i
		sheet.cell(row=RR, column=14).border = thin_border
		sheet.cell(row=RR, column=15).value =dict3[i]
		sheet.cell(row=RR, column=15).border = thin_border
		sheet.cell(row=RR, column=16).value =dict5[dict3[i]]
		sheet.cell(row=RR, column=16).border = thin_border
		# keep tracing no rows
		RR+=1
	# for negative vlue of i
		sheet.cell(row=RR, column=14).value =-1*i
		sheet.cell(row=RR, column=14).border = thin_border
		sheet.cell(row=RR, column=15).value =dict3[-1*i]
		sheet.cell(row=RR, column=15).border = thin_border
		sheet.cell(row=RR, column=16).value =dict5[dict3[-1*i]]
		sheet.cell(row=RR, column=16).border = thin_border
	# keep tracing no rows
		RR+=1
    
	w=2
	sheet.cell(row=1, column=11).value ="octant"
	for i in df["octant"]:
		sheet.cell(row=w, column=11).value =i
		w+=1
	# ---------------------------------------------------------transition count part starts here-------------------------------------------------------
	k=mod
	leng=df['octant'].count()
	try:
		if k<leng:
			zo=k
		else:
			print("mod is too larg!")
	except:
		print("mod is not valid")
		exit()

																	# printing the mod value
	print("mod"+str(k))
	print(leng) 
	z=int(leng/k)+1

	# no rows in output
	R = int(z)
	# no column is constant equal to 9
	C = int(9)
	
	# Initialize matrix
	matrix = []
	number=[1,-1,2,-2,3,-3,4,-4]
	

	zz=int(0)
	coli=35
	rowi=2
	sheet.cell(row=rowi, column=coli).value ="Overall Transition Count"
	rowi+=1
	sheet.cell(row=rowi, column=coli).value =" "
	sheet.cell(row=rowi+2, column=coli-1).value ="from"
	sheet.cell(row=rowi, column=coli+1).value ="to"
	sheet.cell(row=rowi, column=coli+2).value =" "
	sheet.cell(row=rowi, column=coli).border = thin_border
	rowi+=1


	# making our desire desired heading
	sheet.cell(row=rowi, column=coli).value ="count"
	sheet.cell(row=rowi, column=coli).border = thin_border
	sheet.cell(row=rowi, column=coli+1).value ='1'
	sheet.cell(row=rowi, column=coli+1).border = thin_border
	sheet.cell(row=rowi, column=coli+2).value ='-1'
	sheet.cell(row=rowi, column=coli+2).border = thin_border
	sheet.cell(row=rowi, column=coli+3).value ='2'
	sheet.cell(row=rowi, column=coli+3).border = thin_border
	sheet.cell(row=rowi, column=coli+4).value ='-2'
	sheet.cell(row=rowi, column=coli+4).border = thin_border
	sheet.cell(row=rowi, column=coli+5).value ='3'
	sheet.cell(row=rowi, column=coli+5).border = thin_border
	sheet.cell(row=rowi, column=coli+6).value ='-3'
	sheet.cell(row=rowi, column=coli+6).border = thin_border
	sheet.cell(row=rowi, column=coli+7).value ='4'
	sheet.cell(row=rowi, column=coli+7).border = thin_border
	sheet.cell(row=rowi, column=coli+8).value ='-4'
	sheet.cell(row=rowi, column=coli+8).border = thin_border
	rowi+=1

	# below code run for j=1,2,3,4
	# below code is for over all
	qwe=5
	qwec=5
	qwer=0
	for j in range(1,5):
		q=[]
		# below value are transtion count for j
		a=b=c=d=e=f=g=h=0
		# below value are transtion count for -j
		aa=bb=cc=dd=ee=ff=gg=hh=0
		# below cord for countin the all the transtions in the total length of octant
		dd=0
		for k in range(0,leng-1):
			# below cord for tanstion of j to 1 herw j may be 1,2,3,4
			if(df['octant'][k]==j and df['octant'][k+1]==1):
				a=a+1
			# below cord for tanstion of j to -1 herw j may be 1,2,3,4 similerly all the below codes
			if(df['octant'][k]==j and df['octant'][k+1]==-1):
				b=b+1
			if(df['octant'][k]==j and df['octant'][k+1]==2):
				c=c+1
			if(df['octant'][k]==j and df['octant'][k+1]==-2):
				d=d+1
			if(df['octant'][k]==j and df['octant'][k+1]==3):
				e=e+1
			if(df['octant'][k]==j and df['octant'][k+1]==-3):
				f=f+1
			if(df['octant'][k]==j and df['octant'][k+1]==4):
				g=g+1
			if(df['octant'][k]==j and df['octant'][k+1]==-4):
				h=h+1
			if(df['octant'][k]==-j and df['octant'][k+1]==1):
				aa=aa+1
			if(df['octant'][k]==-j and df['octant'][k+1]==-1):
				bb=bb+1
			if(df['octant'][k]==-j and df['octant'][k+1]==2):
				cc=cc+1
			if(df['octant'][k]==-j and df['octant'][k+1]==-2):
				dd=dd+1
			if(df['octant'][k]==-j and df['octant'][k+1]==3):
				ee=ee+1
			if(df['octant'][k]==-j and df['octant'][k+1]==-3):
				ff=ff+1
			if(df['octant'][k]==-j and df['octant'][k+1]==4):
				gg=gg+1
			if(df['octant'][k]==-j and df['octant'][k+1]==-4):
				hh=hh+1   

		sheet.cell(row=rowi, column=coli).value =str(j)
		sheet.cell(row=rowi,column=coli+qwer+1).fill = PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")
		qwer+=1
		sheet.cell(row=rowi, column=coli).border = thin_border
		sheet.cell(row=rowi, column=coli+1).value =str(a)
		sheet.cell(row=rowi, column=coli+1).border = thin_border
		sheet.cell(row=rowi, column=coli+2).value =str(b)
		sheet.cell(row=rowi, column=coli+2).border = thin_border
		sheet.cell(row=rowi, column=coli+3).value =str(c)
		sheet.cell(row=rowi, column=coli+3).border = thin_border
		sheet.cell(row=rowi, column=coli+4).value =str(d)
		sheet.cell(row=rowi, column=coli+4).border = thin_border
		sheet.cell(row=rowi, column=coli+5).value =str(e)
		sheet.cell(row=rowi, column=coli+5).border = thin_border
		sheet.cell(row=rowi, column=coli+6).value =str(f)
		sheet.cell(row=rowi, column=coli+6).border = thin_border
		sheet.cell(row=rowi, column=coli+7).value =str(g)
		sheet.cell(row=rowi, column=coli+7).border = thin_border
		sheet.cell(row=rowi, column=coli+8).value =str(h)
		sheet.cell(row=rowi, column=coli+8).border = thin_border
		rowi+=1
		# adding the transtion counts in new matrix as a row for j to 1,-1,2,-2 and so on

		sheet.cell(row=rowi, column=coli).value =str(-j)
		sheet.cell(row=rowi,column=coli+qwer+1).fill = PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")
		qwer+=1
		sheet.cell(row=rowi, column=coli).border = thin_border
		sheet.cell(row=rowi, column=coli+1).value =str(aa)
		sheet.cell(row=rowi, column=coli+1).border = thin_border
		sheet.cell(row=rowi, column=coli+2).value =str(bb)
		sheet.cell(row=rowi, column=coli+2).border = thin_border
		sheet.cell(row=rowi, column=coli+3).value =str(cc)
		sheet.cell(row=rowi, column=coli+3).border = thin_border
		sheet.cell(row=rowi, column=coli+4).value =str(dd)
		sheet.cell(row=rowi, column=coli+4).border = thin_border
		sheet.cell(row=rowi, column=coli+5).value =str(ee)
		sheet.cell(row=rowi, column=coli+5).border = thin_border
		sheet.cell(row=rowi, column=coli+6).value =str(ff)
		sheet.cell(row=rowi, column=coli+6).border = thin_border
		sheet.cell(row=rowi, column=coli+7).value =str(gg)
		sheet.cell(row=rowi, column=coli+7).border = thin_border
		sheet.cell(row=rowi, column=coli+8).value =str(hh)
		sheet.cell(row=rowi, column=coli+8).border = thin_border
		# adding the transtion counts in new matrix as a row for -j to 1,-1,2,-2 and so on
		rowi+=1

		# -----------------------------------------------------------------------------------------------------------------------------
	
	for i in range(R):          # A for loop for row entries
			# creating a list
			
				# condition for last row of output to cheak the  length of input so it stop when input length reached
			if(i==R-1):
					
				
					sheet.cell(row=rowi, column=coli).value ="Transition Count"
					rowi+=1
					sheet.cell(row=rowi, column=coli).value =str(i*zo)
					sheet.cell(row=rowi+2, column=coli-1).value ="from"
					sheet.cell(row=rowi, column=coli+1).value ="to"
					sheet.cell(row=rowi, column=coli+2).value =str(leng-1)
					sheet.cell(row=rowi, column=coli).border = thin_border
					rowi+=1
						
			else:    
					# appending the range in which we calculating the transtions count  
					
					sheet.cell(row=rowi, column=coli).value ="Transition Count"
					rowi+=1
					sheet.cell(row=rowi, column=coli).value =str(i*zo)
					sheet.cell(row=rowi+2, column=coli-1).value ="from"
					sheet.cell(row=rowi, column=coli+1).value ="to"
					sheet.cell(row=rowi, column=coli+2).value =str((i+1)*zo-1)
					sheet.cell(row=rowi, column=coli).border = thin_border
					rowi+=1

			# making our desire desired heading
			sheet.cell(row=rowi, column=coli).value ="count"
			sheet.cell(row=rowi, column=coli).border = thin_border
			sheet.cell(row=rowi, column=coli+1).value ='1'
			sheet.cell(row=rowi, column=coli+1).border = thin_border
			sheet.cell(row=rowi, column=coli+2).value ='-1'
			sheet.cell(row=rowi, column=coli+2).border = thin_border
			sheet.cell(row=rowi, column=coli+3).value ='2'
			sheet.cell(row=rowi, column=coli+3).border = thin_border
			sheet.cell(row=rowi, column=coli+4).value ='-2'
			sheet.cell(row=rowi, column=coli+4).border = thin_border
			sheet.cell(row=rowi, column=coli+5).value ='3'
			sheet.cell(row=rowi, column=coli+5).border = thin_border
			sheet.cell(row=rowi, column=coli+6).value ='-3'
			sheet.cell(row=rowi, column=coli+6).border = thin_border
			sheet.cell(row=rowi, column=coli+7).value ='4'
			sheet.cell(row=rowi, column=coli+7).border = thin_border
			sheet.cell(row=rowi, column=coli+8).value ='-4'
			sheet.cell(row=rowi, column=coli+8).border = thin_border
			rowi+=1
			qwer=0
			for j in range(1,5):
							
							a=b=c=d=e=f=g=h=0
							aa=bb=cc=dd=ee=ff=gg=hh=0
							for k in range(i*zo,zo*(i+1)):
								# checking wether the k is not greater then no of row present in data frame
								if k>leng-2:
									break
								# below code is same as described in above 
								if(df['octant'][k]==j and df['octant'][k+1]==1):
									a=a+1
								# below cord for tanstion of j to -1 herw j may be 1,2,3,4 similerly all the below codes
								if(df['octant'][k]==j and df['octant'][k+1]==-1):
									b=b+1
								if(df['octant'][k]==j and df['octant'][k+1]==2):
									c=c+1
								if(df['octant'][k]==j and df['octant'][k+1]==-2):
									d=d+1
								if(df['octant'][k]==j and df['octant'][k+1]==3):
									e=e+1
								if(df['octant'][k]==j and df['octant'][k+1]==-3):
									f=f+1
								if(df['octant'][k]==j and df['octant'][k+1]==4):
									g=g+1
								if(df['octant'][k]==j and df['octant'][k+1]==-4):
									h=h+1
								if(df['octant'][k]==-j and df['octant'][k+1]==1):
									aa=aa+1
								if(df['octant'][k]==-j and df['octant'][k+1]==-1):
									bb=bb+1
								if(df['octant'][k]==-j and df['octant'][k+1]==2):
									cc=cc+1
								if(df['octant'][k]==-j and df['octant'][k+1]==-2):
									dd=dd+1
								if(df['octant'][k]==-j and df['octant'][k+1]==3):
									ee=ee+1
								if(df['octant'][k]==-j and df['octant'][k+1]==-3):
									ff=ff+1
								if(df['octant'][k]==-j and df['octant'][k+1]==4):
									gg=gg+1
								if(df['octant'][k]==-j and df['octant'][k+1]==-4):
									hh=hh+1   
							
							sheet.cell(row=rowi, column=coli).value =str(j)
							sheet.cell(row=rowi,column=coli+qwer+1).fill = PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")
							qwer+=1
							sheet.cell(row=rowi, column=coli).border = thin_border
							sheet.cell(row=rowi, column=coli+1).value =str(a)
							sheet.cell(row=rowi, column=coli+1).border = thin_border
							sheet.cell(row=rowi, column=coli+2).value =str(b)
							sheet.cell(row=rowi, column=coli+2).border = thin_border
							sheet.cell(row=rowi, column=coli+3).value =str(c)
							sheet.cell(row=rowi, column=coli+3).border = thin_border
							sheet.cell(row=rowi, column=coli+4).value =str(d)
							sheet.cell(row=rowi, column=coli+4).border = thin_border
							sheet.cell(row=rowi, column=coli+5).value =str(e)
							sheet.cell(row=rowi, column=coli+5).border = thin_border
							sheet.cell(row=rowi, column=coli+6).value =str(f)
							sheet.cell(row=rowi, column=coli+6).border = thin_border
							sheet.cell(row=rowi, column=coli+7).value =str(g)
							sheet.cell(row=rowi, column=coli+7).border = thin_border
							sheet.cell(row=rowi, column=coli+8).value =str(h)
							sheet.cell(row=rowi, column=coli+8).border = thin_border
							rowi+=1
							# adding the transtion counts in new matrix as a row for j to 1,-1,2,-2 and so on

							sheet.cell(row=rowi, column=coli).value =str(-j)
							sheet.cell(row=rowi, column=coli).border = thin_border
							sheet.cell(row=rowi,column=coli+qwer+1).fill = PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")
							qwer+=1
							sheet.cell(row=rowi, column=coli+1).value =str(aa)
							sheet.cell(row=rowi, column=coli+1).border = thin_border
							sheet.cell(row=rowi, column=coli+2).value =str(bb)
							sheet.cell(row=rowi, column=coli+2).border = thin_border
							sheet.cell(row=rowi, column=coli+3).value =str(cc)
							sheet.cell(row=rowi, column=coli+3).border = thin_border
							sheet.cell(row=rowi, column=coli+4).value =str(dd)
							sheet.cell(row=rowi, column=coli+4).border = thin_border
							sheet.cell(row=rowi, column=coli+5).value =str(ee)
							sheet.cell(row=rowi, column=coli+5).border = thin_border
							sheet.cell(row=rowi, column=coli+6).value =str(ff)
							sheet.cell(row=rowi, column=coli+6).border = thin_border
							sheet.cell(row=rowi, column=coli+7).value =str(gg)
							sheet.cell(row=rowi, column=coli+7).border = thin_border
							sheet.cell(row=rowi, column=coli+8).value =str(hh)
							sheet.cell(row=rowi, column=coli+8).border = thin_border
							# adding the transtion counts in new matrix as a row for -j to 1,-1,2,-2 and so on
							rowi+=1
															
			rowi+=2
	rowi=3
	# -----------------------------------------------------------longest subsequence part--------------------------------------------------------------
# -----------------------------------------------------------------longest subsequence heading -------------------------------------------------------------------
	coli+=10
	sheet.cell(row=1, column=coli).value ="Longest Subsquence Length"
	sheet.cell(row=1, column=coli).border = thin_border
	# making our desire matrix having desired heading
	sheet.cell(row=2, column=coli).value ="Count"
	sheet.cell(row=2, column=coli).border = thin_border

	sheet.cell(row=2, column=coli+1).value ="Longest Subsquence Length"
	sheet.cell(row=2, column=coli+1).border = thin_border

	sheet.cell(row=2, column=coli+2).value ="Count"
	sheet.cell(row=2, column=coli+2).border = thin_border
	# making our desire matrix having desired heading
	sheet.cell(row=2, column=coli+5).value ="Count"
	sheet.cell(row=2, column=coli+5).border = thin_border

	sheet.cell(row=2, column=coli+6).value ="Longest Subsquence Length"
	sheet.cell(row=2, column=coli+6).border = thin_border

	sheet.cell(row=2, column=coli+7).value ="Count"
	sheet.cell(row=2, column=coli+7).border = thin_border
	# ---------------------------------------------------below code is for calculating the longest subsequence ans its count---------------------
	countp=1
	countn=-1

	for i in range(3,11):
		
		for j in range(0,1):
			# for the +1,+2,+3,+4
			if j==0 and i%2==0:
				sheet.cell(row=i, column=j+coli).value =countp
				sheet.cell(row=i, column=j+coli).border = thin_border
				
				maxo=0
				count=0
				t=0
				for k in df['octant']:
						if k==countp:
							t=t+1
						else:
							t=0
						if(maxo<t):
							maxo=t
						
						
				for k in df['octant']:
						if k==countp:
							t=t+1
						else:
							t=0
						if(maxo==t):
							count=count+1
						
				# adding the longest subsequece in the matrix
				sheet.cell(row=i, column=j+1+coli).value =maxo
				# adding the border to it
				sheet.cell(row=i, column=j+1+coli).border = thin_border
				# no times the longest susequence appears in the octant matrix
				sheet.cell(row=i, column=j+2+coli).value =count
				# adding the border to it
				sheet.cell(row=i, column=j+2+coli).border = thin_border
				countp=countp+1
			# for the -1,-2,-3,-4  
			if j==0 and i%2!=0:
				sheet.cell(row=i, column=j+coli).value =countn
				sheet.cell(row=i, column=j+coli).border = thin_border
				maxo=0
				count=0
				t=0
				for k in df['octant']:
						if k==countn:
							t=t+1
						else:
							t=0
						if(maxo<t):
							maxo=t
						
						
				for k in df['octant']:
						if k==countn:
							t=t+1
						else:
							t=0
						if(maxo==t):
							count=count+1
						
				# adding the longest subsequece in the matrix
				sheet.cell(row=i, column=j+1+coli).value =maxo
				# adding the border to it
				sheet.cell(row=i, column=j+1+coli).border = thin_border
					# no times the longest susequence appears in the octant matrix
				sheet.cell(row=i, column=j+2+coli).value =count
				# adding the border to it
				sheet.cell(row=i, column=j+2+coli).border = thin_border
				
				
				countn=countn-1  
	# -------------------below code for calculating the longest subsequence ans its count  with time included--------------------------------------------------------

	#  making a new variable to keep track of column so that i can easly add new columns
	star=3
	sheet.cell(row=1, column=coli+5).value ="Longest Subsquence Length with Range"
	sheet.cell(row=1, column=coli+5).border = thin_border
	print("line 820")
	for i in range(1,5):
		
		for j in range(5,6):
			# for the +1,+2,+3,+4
			if j==5 :

				sheet.cell(row=star, column=j+coli).value =1*i
				sheet.cell(row=star, column=j+coli).border = thin_border
				
				maxo=0
				count=0
				t=0
				# below code for calulating the max subsequence length
				for k in df['octant']:
						if k==i:
							t=t+1
						else:
							t=0

						if(maxo<t):
							maxo=t
						
				# below code for calulating the no frequency for occurence max subsequence length   
				for k in df['octant']:
						if k==i:
							t=t+1
						else:
							t=0
						if(maxo==t):
							count=count+1
						
				# adding the longest subsequece in the matrix
				sheet.cell(row=star, column=j+1+coli).value =maxo
				# adding the border to it
				sheet.cell(row=star, column=j+1+coli).border = thin_border
				# no times the longest susequence appears in the octant matrix
				sheet.cell(row=star, column=j+2+coli).value =count
				# adding the border to it
				sheet.cell(row=star, column=j+2+coli).border = thin_border
				# increasing the value of varble to keep track of no. columns
				star=star+1

				# adding the new row 
				# for time to from.
				sheet.cell(row=star, column=j+coli).value ="Time"
				# adding the border to it
				sheet.cell(row=star, column=j+coli).border = thin_border
				sheet.cell(row=star, column=j+1+coli).value ="From"
				# adding the border to it
				sheet.cell(row=star, column=j+1+coli).border = thin_border
				sheet.cell(row=star, column=j+2+coli).value ="To"
				# adding the border to it
				sheet.cell(row=star, column=j+2+coli).border = thin_border
					# increasing the value of varble to keep track of no. columns
				star=star+1

				# below code for finding where the max subsequnce length starts and ends.
				# below variable to keep track no itrations whilch halps to find the dirst occurence of an element ifrom the max subsequence length
				qq=0
				for k in df['octant']:
						qq=qq+1
						if k==i:
							t=t+1
						else:
							t=0
						if(maxo==t):
							sheet.cell(row=star, column=j+coli).value =" "
							# adding the border to it
							sheet.cell(row=star, column=j+coli).border = thin_border
							try:
								sheet.cell(row=star, column=j+1+coli).value =df['T'][qq-maxo]
								# adding the border to it
								sheet.cell(row=star, column=j+1+coli).border = thin_border
								sheet.cell(row=star, column=j+2+coli).value =df['T'][qq-1]
							except:
								print("error in  line 899")
							# adding the border to it
							sheet.cell(row=star, column=j+2+coli).border = thin_border
							# increasing the value of varble to keep track of no columns
							star=star+1			
			# same code as above only for negatve of i or(if above is for 1 than this below code is for -1 and so on).
			if j==5 :
				i=-1*i
				sheet.cell(row=star, column=j+coli).value =1*i
				sheet.cell(row=star, column=j+coli).border = thin_border				
				maxo=0
				count=0
				t=0
				for k in df['octant']:
						if k==i:
							t=t+1
						else:
							t=0

						if(maxo<t):
							maxo=t	
				for k in df['octant']:
						if k==i:
							t=t+1
						else:
							t=0
						if(maxo==t):
							count=count+1
						
				# adding the longest subsequece in the matrix
				sheet.cell(row=star, column=j+1+coli).value =maxo
				# adding the border to it
				sheet.cell(row=star, column=j+1+coli).border = thin_border
				# no times the longest susequence appears in the octant matrix
				sheet.cell(row=star, column=j+2+coli).value =count
				# adding the border to it
				sheet.cell(row=star, column=j+2+coli).border = thin_border
					# increasing the value of varble to keep track of no columns
				star=star+1
				# adding the new row 
				# for time to from.
				sheet.cell(row=star, column=j+coli).value ="Time"
				# adding the border to it
				sheet.cell(row=star, column=j+coli).border = thin_border
				sheet.cell(row=star, column=j+1+coli).value ="From"
				sheet.cell(row=star, column=j+1+coli).border = thin_border
				sheet.cell(row=star, column=j+2+coli).value ="To"
				sheet.cell(row=star, column=j+2+coli).border = thin_border
				# increasing the value of varble to keep track of no columns
				star=star+1
				# below code for finding where the max subsequnce length starts and ends.
				# below variable to keep track no itrations whilch halps to find the dirst occurence of an element ifrom the max subsequence length
				qq=0
				for k in df['octant']:
						qq=qq+1
						if k==i:
							t=t+1
						else:
							t=0
						if(maxo==t):
							sheet.cell(row=star, column=j+coli).value =" "
							# adding the border to it
							sheet.cell(row=star, column=j+coli).border = thin_border
							# sheet.cell(row=star, column=j+1+coli).value =df['Time'][qq-maxo]
							# # adding the border to it
							# sheet.cell(row=star, column=j+1+coli).border = thin_border
							# sheet.cell(row=star, column=j+2+coli).value =df['Time'][qq-1]
							try:
								sheet.cell(row=star, column=j+1+coli).value =df['T'][qq-maxo]
								# adding the border to it
								sheet.cell(row=star, column=j+1+coli).border = thin_border
								sheet.cell(row=star, column=j+2+coli).value =df['T'][qq-1]
							except:
								print("error in  line 899")
							# adding the border to it
							sheet.cell(row=star, column=j+2+coli).border = thin_border
							star=star+1							
		i=-1*i
	# finding the name output file using input file name 
	qwes=file[6:]
	print(qwes)
	book.save("output\\"+qwes[0:len(qwes) - 5]+"cm_vel_octant_analysis_mod_"+str(mod)+".xlsx")
	print("sucssesfully done!")
	
# finding the total time taken	 
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
